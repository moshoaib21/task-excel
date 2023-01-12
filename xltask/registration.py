

from django.http import response , HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth import authenticate, login ,logout
from django.conf import settings
from django.core.mail import send_mail
from django.contrib.auth.decorators import login_required
from openpyxl import*
import openpyxl
from .models import *
import os

def register(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        firstname = request.POST.get('firstname')
        lastname = request.POST.get('lastname')
        password = request.POST.get('pass1')
        password2 = request.POST.get('pass2')

        myuser = User.objects.create_user(username, email, password)
        myuser.first_name = firstname
        myuser.last_name = lastname
        myuser.save()
        message="Your registration is successfully Now you can login "
        subject=f" {firstname} Welcome to xltask app"
        recipient_list=[email]
        email_from=settings.EMAIL_HOST_USER
        send_mail(subject,message,email_from,recipient_list)

        messages.success(request, "Your account has been created")
        return redirect('signin')
    return render(request,'signup.html')


def signin(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('pass1')
        
        user = authenticate(username = username, password = password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        
        else:
            messages.error(request, "Bad request")
            return redirect('/')
    return render(request, 'signin.html')


def signout(request):
    logout(request)
    messages.success(request, "You have been logged out")
    return redirect('/')

def start(request):
    return render(request, 'home.html')

@login_required()
def dashboard(request):
    return render(request, 'dashboard.html')



def Excel_store(request):

    response = HttpResponse(content_type='aplication/vnd.ms-excel')
    response['content-Disposition'] = 'attachment ; filename="EmployeeAttendance-store.xlsx"'
    row_num1=0
    wb = Workbook()
    sheet=Workbook.active
   
    rows= Employee.objects.all().values_list( 'department','firstname','lastname','gender','address','salary').order_by('-firstname')
  
   
    ws = wb.active
  
  
    ws.append( ['department','firstname','lastname','gender','address','salary'])
    # Rows can also be appended
    for row in rows:
        row_num1 += 1
        z=[i for i in row]
        ws.append(z)
  
    
    wb.save(response)

    return response


     
def save_data(request):
    lst=[]
   
    file=request.FILES["file"]
    obj= xl.objects.create( file=file)
    path=str(obj.file)

    
    dataframe = openpyxl.load_workbook(path)
    
    wb=dataframe.active
    # for row in range(1, wb.max_row):
    
    #     for col in range(1, wb.max_column):
            
    for row in wb.values:
        lst.append(row)
        
       
    lst.pop(0)
    for i in lst:
        department=i[0]
        firstname=i[1]
        lastname=i[2]
        gender=i[3]
        address=i[4]
        salary=i[5]
        obj=Employee.objects.create(department_id=department,firstname=firstname,lastname=lastname,gender=gender,address=address,salary=salary)
        obj.save()
   
   
    
    
    return render(request,'dashboard.html')
    






# def upload_file():
    # Get the file data from the POST request

    # file_data = request.files['file']
    # print("filename.......",file_data)

    # file_data.save(os.path.join('static', secure_filename(file_data.filename)))

    # wb = load_workbook(file_data)
    # sheet = wb.active
    # conn = sqlite3.connect('mydatabase.db')
    # cursor = conn.cursor()

    # cursor.execute('''CREATE TABLE IF NOT EXISTS data (did INTEGER, data TEXT)''')
    # for row in sheet.iter_rows(min_row=2):
    #     did = row[0].value
    #     data = row[1].value
    #     cursor.execute('''INSERT OR IGNORE INTO data (did, data) VALUES (?, ?)''', (did, data))

    # conn.commit()
    # conn.close()
    # return jsonify({'message': 'addded in db'})
